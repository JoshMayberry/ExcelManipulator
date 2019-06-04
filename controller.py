__version__ = "2.5.0"
##Does not use win32com

#Import standard elements
import os
import sys
import copy
import warnings
import subprocess
import collections
from ctypes import windll as ctypesWindll #Used to determine screen dpi

#Import the openpyxl module to work with excel sheets
import openpyxl

#Import needed support modules
import PIL.ImageGrab

import MyUtilities.common
NULL_private = MyUtilities.common.Singleton("NULL", state = False, private = True)

#Controllers
def build(*args, **kwargs):
	"""Starts the GUI making process."""

	return Excel(*args, **kwargs)

#Required Modules
##py -m pip install
	# openpyxl

class Utilities(MyUtilities.common.Container, MyUtilities.common.CommonFunctions, MyUtilities.common.EnsureFunctions):
	def __init__(self, child_class = None):

		#Initialize Inherited Modules
		MyUtilities.common.Container.__init__(self)
		MyUtilities.common.CommonFunctions.__init__(self)
		MyUtilities.common.EnsureFunctions.__init__(self)

		#Internal Variables
		self.child_class = child_class
		if (child_class is not None):
			self.current = None
			self.child_uniqueName = f"{self.child_class.__name__}_{{}}"

	def __contains__(self, key):
		return self._get(key, returnExists = True)

	def __getitem__(self, key):
		return self.getChild(label = key)

	def __str__(self):
		"""Gives diagnostic information on this when it is printed out."""

		output = MyUtilities.common.Container.__str__(self)
		output += f"-- Title: {self.title}\n"
		return output

	def getChild(self, label = None, *args, **kwargs):
		"""Returns an child.

		label (str) - What the child is called
			- If None: Returns the current child
			- If does not exist: Creates a new child

		Example Input: getChild()
		"""

		if (self.child_class is None):
			raise NotImplementedError()

		if (label is None):
			if (self.current is not None):
				return self.current
			label = self.getUnique(self.child_uniqueName)
			select = True
		else:
			select = False

		child = self._get(label, returnForNone = None)
		if (child is None):
			child = self.new(*args, label = label, **kwargs)
		if (select):
			self.select(child)
		return child

	def select(self, label, thing = None):
		"""Selects the given child as the current one.

		Example Input: select("lorem")
		"""

		if (thing is not None):
			for item in self:
				if (item.thing == thing):
					self.current = item
					return
			raise NotImplementedError()

		if (self.child_class is None):
			raise NotImplementedError()
		if (label is None):
			raise NotImplementedError()

		if (isinstance(label, self.child_class)):
			self.current = label
		else:
			self.current = self[label]
		return self.current

	def new(self, *args, label = None, **kwargs):
		"""Creates a new child and saves it in memmory.

		label (str) - The label of the child
		firstSheet (str) - The label for the first sheet in the child
			- If None: The child will start off without any sheets

		Example Input: new(label = "test")
		"""

		if (label is None):
			label = self.getUnique(self.child_uniqueName)

		child = self.child_class(self, label, *args, **kwargs)
		self[label] = child
		return child

	@MyUtilities.common.makeProperty(default = None)
	class title():
		"""The title of 'thing'.
			- If None: Will use 'label'
		"""

		def setter(self, value):
			self.thing.title = f"{self.ensure_default(value, default = self.label)}"

		def getter(self):
			return self.thing.title

	@classmethod
	def convertColumn(cls, column, row = None):
		"""Converts a column number to a column letter, and returns it to the user as a string.

		column (int)  - The index of the column

		Example Input: convertColumn(3)
		"""

		#Convert Column if needed
		if (isinstance(column, int)):
			#Check for past Z
			count = 0
			bonusColumn = ""
			while True:
				count += 1
				#Does the ascii letter go past Z? If so, create addition letter
				if (openpyxl.utils.get_column_letter(count).isupper()):
					break
				else:
					column -= 26
					bonusColumn = openpyxl.utils.get_column_letter(count)

			#Set new Column
			column = bonusColumn + openpyxl.utils.get_column_letter(column)

		if (row is None):
			return column
		return f"{column}{row}"

class Excel(Utilities):
	def __init__(self):
		"""Works with excel files.
		Documentation for openpyxl can be found at: https://openpyxl.readthedocs.io/en/default/index.html

		Example Input: Excel()
		"""

		super().__init__(child_class = self.Book)

	def save(self, *args, label = None, **kwargs):
		"""Saves the workbook to a specified location.

		Example Input: save(label = "test")
		"""

		book = self[label]
		self[label].save(*args, **kwargs)
		return book

	def load(self, *args, label = None, **kwargs):
		"""Loads a workbook from a specified location into memmory.

		Example Input: load(label = "test")
		"""

		book = self[label]
		book.load(*args, **kwargs)
		return book

	def run(self, *args, label = None, **kwargs):
		"""Opens the excel file for the user.

		Example Input: run(label = "converted")
		"""

		book = self[label]
		book.run(*args, **kwargs)
		return book

	class Book(Utilities):
		def __init__(self, parent, label, *, firstSheet = None, title = None, filePath = None, readOnly = False, writeOnly = False):
			"""A handle for the workbook.
			See: https://openpyxl.readthedocs.io/en/latest/optimized.html
			See: https://stackoverflow.com/questions/21875249/memory-error-using-openpyxl-and-large-data-excels/21875423#21875423

			firstSheet (str) - The label for the first sheet in the workbook
				- If None: The workbook will start off without any sheets

			Example Input: Book(self, label)
			Example Input: Book(self, label, firstSheet = "Sheet_1")
			"""

			#Initialize Inherited Modules
			super().__init__(child_class = self.Sheet)

			#Internal Variables
			self.label = label
			self.parent = parent
			self.filePath = filePath
			self.firstSheet = firstSheet
			self.readOnly = readOnly
			
			self.imageCatalogue = collections.defaultdict(dict) #Used to catalogue all of the images in the document. {sheet title: {top-left corner cell (row, column): image as a PIL image}}

			self.thing = openpyxl.Workbook(write_only = writeOnly)
			self.title = title

			if (firstSheet != None):
				sheet = self[firstSheet]
				sheet.select()

		# def __exit__(self, exc_type, exc_value, traceback):
		# 	if (traceback is None):
		# 		self.save()
		# 	return super().__exit__(exc_type, exc_value, traceback)

		@MyUtilities.common.makeProperty(default = None)
		class filePath():
			def setter(self, value):
				self._filePath = self.ensure_filePath(self.ensure_default(value, default = self.label), ending = (".xls", ".xlsx"), checkExists = False)

			def getter(self):
				return self._filePath

		def add(self, label = None, *, changeToSheet = True, **kwargs):
			"""Adds a new sheet to the excel file.

			position (int)       - Where to insert the sheet at
				- If None: Insert at the end
			label (str)          - The name of the sheet
				- If None: It is given the default name (ie: Sheet, Sheet1, Sheet2, etc.)
			tabColor (str)       - The RRGGBB color code for the tab
				- If None: it is the default white 
			changeToSheet (bool) - Wether to change the current sheet to this new sheet or not

			Example Input: add()
			Example Input: add("Sheet1", position = 0, tabColor = "1072BA")
			"""

			sheet = self.getChild(label = label, **kwargs)
			if (changeToSheet):
				self.select(label)

			return sheet

		def remove(self, label):
			"""Removes a sheet from the book.

			label (str) - The name of the sheet to be removed from the book

			Example Input: remove("sheet1")
			"""

			self[label].remove()

		def _get(self, label = None, *, returnExists = False, returnForNone = NULL_private, **kwargs):
			"""Overridden to account for being given a sheet's title or position.

			Example Input: _get()
			Example Input: _get("Sheet1")
			"""

			def getThing():
				nonlocal self, label

				if (label is None):
					return self.thing.active

				if (isinstance(label, str)):
					if (label in self.thing):
						return self.thing[label]
					return

				if (label in self.thing.worksheets):
					return self.thing.worksheets[label]

			def getChild():
				thing = getThing()
				for item in self._dataCatalogue.values():
					if (item.thing is thing):
						return item

			##################################

			if (returnExists):
				if (super()._get(label = label, returnExists = True, **kwargs)):
					return True
			else:
				answer = super()._get(label = label, returnForNone = None, **kwargs) or None
				if (answer is not None):
					return answer

			answer = getChild()
			if (returnExists):
				return answer is not None

			if (answer is not None):
				return answer

			if (returnForNone is not NULL_private):
				return returnForNone

			errorMessage = f"{sheet} is not nested in {self.__repr__()}"
			raise KeyError(errorMessage)

		def getAllSheetNames(self):
			"""Returns a list of all the sheet names as strings."""

			return self.thing.get_sheet_names()

		def _getFilePath(self, filePath = None, temporary = False):
			if (filePath is None):
				_filePath = self.filePath
			else:
				_filePath = self.ensure_filePath(filePath, ending = (".xls", ".xlsx"), checkExists = False)

			if (not temporary):
				return _filePath
			return "{}_temp{}".format(*os.path.splitext(_filePath))

		def save(self, filePath = None, overlayOk = True, temporary = False, saveImages = True):
			"""Saves the workbook to a specified location.

			filePath (str)   - Where the file is located
			overlayOk (bool) - If True: Images can overlap. If False: Any images under otehr ones will be deleted. If None: Images will be scooted to the right until they are ont under another
			temporary (bool) - If True: The file will be saved under the same name, but with "_temp" after it. For debugging things
			saveImages (bool) - If True: Images in the document will be preserved upon loading
				Images, charts, etc. are not read by openpyxl.
				In order to preserve images, charts, etc., each image is loaded and re-written into the loaded workbook
				Method for preservation from http://www.penwatch.net/cms/?p=582
				Help from: code.activestate.com/recipes/528870-class-for-writing-content-to-excel-and-formatting

			Example Input: save()
			"""

			try:
				self.thing.save(self._getFilePath(filePath = filePath, temporary = temporary))
			except IOError:
				#A book by that name is already open
				print("ERROR: The excel file is still open. The file has still been saved. Just close the current file without saving.")

		def load(self, filePath = None, readImages = False):
			"""Loads a workbook from a specified location into memmory.

			filePath (str) - Where the file is located
			readImages (bool) - If True: Images in the document will be preserved upon loading
				Images, charts, etc. are not read by openpyxl.
				In order to preserve images, charts, etc., each image is loaded and re-written into the loaded workbook
				Method for preservation from http://www.penwatch.net/cms/?p=582
				Help from: code.activestate.com/recipes/528870-class-for-writing-content-to-excel-and-formatting

			Example Input: load()
			"""

			self.thing = openpyxl.load_workbook(self._getFilePath(filePath = filePath), read_only = self.readOnly)
			self._mapWorksheets(readImages = readImages)

		def _mapWorksheets(self, readImages = False):
			if (not self.thing.worksheets):
				return

			for i, sheet in enumerate(self.thing.worksheets):
				self[i] = self.new(label = i, thing = sheet)
			self.select(None, thing = self.thing.active)

		def run(self, filePath = None):
			"""Opens the excel file for the user.

			filePath (str) - Where the file is located

			Example Input: run()
			"""

			fileName = self._getFilePath(filePath = filePath)
			try:
				os.startfile(fileName)
			except AttributeError:
				subprocess.call(['open', fileName])
			
		class Sheet(Utilities):
			def __init__(self, parent, label, position = None, thing = None, title = None, tabColor = None):
				"""A handle for an excel sheet."""
				super(Excel.Book.Sheet, self).__init__()
				
				self.label = label
				self.parent = parent

				if (thing is not None):
					self.thing = thing
				else:
					if ((len(self.parent) is 0) and (self.parent.firstSheet is None)):
						self.thing = self.parent.thing.active
					else:
						self.thing = None
					
					if (self.thing is None):
						if (position is not None):
							self.thing = self.parent.thing.create_sheet(position)
						else:
							self.thing = self.parent.thing.create_sheet()

					self.title = title
					self.tabColor = tabColor

			def getRowCount(self):
				#https://stackoverflow.com/questions/35408339/is-there-any-method-to-get-the-number-of-rows-and-columns-present-in-xlsx-sheet/35408471#35408471
				return self.thing.max_row

			def getColumnCount(self):
				#https://stackoverflow.com/questions/35408339/is-there-any-method-to-get-the-number-of-rows-and-columns-present-in-xlsx-sheet/35408471#35408471
				return self.thing.max_column

			def remove(self):
				"""Removes this sheet from the book.

				Example Input: remove("sheet1")
				"""

				self.parent.thing.remove_sheet(self.thing)
				del self.parent[self.label]

			@MyUtilities.common.makeProperty(default = None)
			class tabColor():
				"""The tab color of the sheet; The RRGGBB color code for the tab.
					- If None: it is the default white
				"""

				def setter(self, tabColor):
					self.thing.sheet_properties.tabColor = tabColor

				def getter(self):
					return self.thing.sheet_properties.tabColor

			def getCell(self, row, column):
				"""Returns a specific cell object.
				The top-left corner is row (1, 1) not (0, 0).

				row (int)    - The index of the row
				column (int) - The index of the column. Can be a char

				Example Input: getCell(1, 2)
				"""
				
				if ((row is None) or (column is None)):
					raise NotImplementedError()

				return self.thing[self.convertColumn(row = row, column = column)]

			def getCellValue(self, row = None, column = None, *, cell = None):
				"""Returns the contents of a cell.
				The top-left corner is row (1, 1) not (0, 0).

				row (int)     - The index of the row
				column (int)  - The index of the column. Can be a char
				cell (object) - An openpyxl cell object

				Example Input: getCellValue(1, 2)
				Example Input: getCellValue(cell = myCell)
				"""

				if (cell is None):
					cell = self.getCell(row = row, column = column)
				return cell.value

			def getCellImage(self, row, column):
				"""Returns a PIL image object from a cell. 
				Returns 'None' if no image was found on the sheet.
				The top-left corner is row (1, 1) not (0, 0).

				row (int)    - The index of the row
				column (int) - The index of the column

				Example Input: getCellImage(1, 2)
				"""

				title = self.title
				if (title not in self.parent.imageCatalogue):
					return

				catalogue = self.parent.imageCatalogue[title]
				if ((row, column) not in catalogue):
					return

				return catalogue[(row, column)]

			#Setters
			def setCell(self, row = None, column = None, value = None, *, cell = None):
				"""Writes the value of a cell.
				The top-left corner is row (1, 1) not (0, 0).

				value (any)    - What will be written to the cell
				row (int)    - The index of the row
				column (int) - The index of the column. Can be a char

				Example Input: setCell(1, 2, 42)
				Example Input: setCell(1, "B", 3.14)
				Example Input: setCell(1, 2, "Hello World")
				"""

				if (cell is None):
					cell = self.getCell(row = row, column = column)

				#Write Value
				for _cell in self.ensure_container(cell):
					_cell.value = f"{value}" #Make sure input is a valid ascii

			def appendRow(self, contents = None):
				"""Appends a row to the end of the file.

				contents (list) - What the cells in the row will contain. If None, the row will be blank

				Example Input: appendRow()
				Example Input: appendRow([0, 1, 2, 3, 4, 5])
				"""

				#Find the last row
				row = len(tuple(self.thing.iter_rows())) + 1

				#Write to cells
				if ((contents != None) and (len(contents) != 0)):
					for column, item in enumerate(contents):
						self.setCell(row, column + 1, item)
				else:
					self.setCell(row, 1, " ")

			def appendColumn(self, contents = None):
				"""Appends a column to the end of the file.

				contents (list) - What the cells in the column will contain. If None, the column will be blank

				Example Input: appendColumn()
				Example Input: appendColumn([0, 1, 2, 3, 4, 5])
				"""

				#Find the last column
				column = len(tuple(self.thing.iter_cols())) + 1

				#Write to cells
				if ((contents != None) and (len(contents) != 0)):
					for row, item in enumerate(contents):
						self.setCell(row + 1, column, item)
				else:
					self.setCell(1, column, " ")

			def setCellFormula(self, row = None, column = None, formula = None, *args, cell = None):
				"""Writes an excel formula value to a cell.
				The top-left corner is row (1, 1) not (0, 0).

				row (int)    - The index of the row
				column (int) - The index of the column. Can be a char
				formula (str)  - The excel formula
				args*          - The arguments for the formula

				Example Input: setCell(1, 2, "SUM", 1, 2, 3, 4, 5, 6, 7)
				"""

				#Ensure formula format
				formula = formula.upper()

				#Check if formula exists
				if (formula not in openpyxl.utils.FORMULAE):
					errorMessage = f"Unknown Formula {formula}"
					raise KeyError(errorMessage)

				if (cell is None):
					cell = self.getCell(row = row, column = column)

				cell.value = f"={formula}{', '.join(f'{item}' for item in args)}"

			def setCellImage(self, row, column, imagePath, xSize = None, ySize = None, keepAspectRatio = True):
				"""Inserts an image to a cell.
				The top-left corner is row (1, 1) not (0, 0).

				imagePath (str) - The path to the image. Can be a PIL image
				row (int)       - The index of the row
				column (int)    - The index of the column. Can be a char
				xSize (int)     - The width (in pixels) of the image on the excel file. If None: Do not size to this
				ySize (int)     - The height (in pixels) of the image on the excel file. If None: Do not Size to this
				keepAspectRatio (bool) - If True: The image aspect ratio will be preserved. The re-size will go off of the largest side

				Example Input: setCellImage(1, 2, "test.jpg")
				Example Input: setCellImage(1, 2, image)
				"""

				column = self.convertColumn(column)

				image = openpyxl.drawing.image.Image(imagePath, size = (xSize, ySize), nochangeaspect = keepAspectRatio)
				self.thing.add_image(image, f"{column}{row}")

				self.parent.imageCatalogue[self.title][(row, column)] = image

			fontStyle_catalogue = MyUtilities.common._dict({
				"default": {"style": "Normal"}, None: "${default}",
				"calibri": {"name": "Calibri", "size": 11, "bold": False, "italic": False, "vertAlign": None, "underline": "none", "strike": False, "color": "FF000000"},
				"title": {"style": "Title"}, 
				"headline1": {"style": "Headline 1"}, 
				"headline2": {"style": "Headline 2"}, 
				"headline3": {"style": "Headline 3"}, 
				"headline4": {"style": "Headline 4"}, 
				"calculation": {"style": "Calculation"}, 
				"warningText": {"style": "Warning Text"}, 
				"input": {"style": "Input"}, 
				"output": {"style": "Output"}, 
				"good": {"style": "Good"}, 
				"bad": {"style": "Bad"}, 
				"neutral": {"style": "Neutral"}, 
			}, caseSensitive = False, typeSensitive = False)
			
			def setCellStyle(self, row, column, *, font = None, bold = None, italic = None, size = None):
				"""Changes the style of the text in a cell.
				The top-left corner is row (1, 1) not (0, 0).

				row (int)    - The index of the row
				column (int) - The index of the column. Can be a char
				font (str)   - The name of the font style. If None: the current font style will be used
				bold (bool)  - If True: bolds the text. If False: unbolds the text. If None: does not change the boldness of the text
				size (int)   - The size of the text in the cell. The size '11' is typical. Can be a string

				Example Input: setCellStyle(2, 3)
				Example Input: setCellStyle(2, 3, font = "calibri")
				Example Input: setCellStyle(2, 3, bold = True)
				"""

				cell = self.getCell(row, column)

				if (font is None):
					font = copy.copy(cell.font)
				elif (font not in self.fontStyle_catalogue):
					raise NotImplementedError(font)
				else:
					font = openpyxl.styles.Font(**self.fontStyle_catalogue[font])

				if (bold is not None):
					font.bold = bold
				if (italic is not None):
					font.italic = italic
				if (size is not None):
					font.size = int(size)

				cell.font = font

			def mergeCells(self, startRow, startColumn, endRow, endColumn):
				"""Merges a range of cells.
				The top-left corner is row (1, 1) not (0, 0).

				startRow (int) - The index of the left end row of the cells to merge. Can be a char
				startColumn (int) - The index of the left end column of the cells to merge
				endRow (int) - The index of the right end row of the cells to merge. Can be a char
				endColumn (int) - The index of the left end column of the cells to merge

				Example Input: mergeCells(1, 2, 3, 2)
				"""

				if (isinstance(startColumn, str)):
					return self.thing.mergeCells(f"{startColumn}{startRow}:{endColumn}{endRow}")

				if (isinstance(endColumn, int)):
					return self.thing.merge_cells(start_row = startRow, start_column = startColumn, end_row = endRow, end_column = endColumn)

				raise NotImplementedError()

			def unmergeCells(self, startRow, startColumn, endRow, endColumn):
				"""Unmerges a range of cells.
				The top-left corner is row (1, 1) not (0, 0).

				startRow (int) - The index of the left end row of the cells to unmerge. Can be a char
				startColumn (int) - The index of the left end column of the cells to unmerge
				endRow (int) - The index of the right end row of the cells to unmerge. Can be a char
				endColumn (int) - The index of the left end column of the cells to unmerge

				Example Input: unmergeCells(1, 2, 3, 2)
				"""

				if (isinstance(startColumn, str)):
					return self.thing.unmergeCells(f"{startColumn}{startRow}:{endColumn}{endRow}")

				if (isinstance(endColumn, int)):
					return self.thing.unmerge_cells(start_row = startRow, start_column = startColumn, end_row = endRow, end_column = endColumn)

				raise NotImplementedError()

			def hideColumns(self, startColumn, endColumn):
				"""Hides a range of columns.

				startColumn (int) - The index of the left end column of the cells to hide. Can be a char
				endColumn (int)   - The index of the right end column of the cells to hide. Can be a char

				Example Input: hideColumns(1, 2)
				Example Input: hideColumns("A", "D")
				"""

				self.thing.column_dimensions.group(self.convertColumn(startColumn), self.convertColumn(endColumn), hidden = True)

			def _filter_trailingNone(self, column):
				column = tuple(column)
				while (column):
					if (self.getCellValue(cell = column[-1]) == None):
						column.pop(-1)
					else:
						break
				return column

			def _yieldContents(self, item, *, useForNone = ""):
				if (useForNone is None):
					for piece in item:
						yield self.getCellValue(cell = piece)
					return

				for piece in item:
					value = self.getCellValue(cell = piece)
					if (value is None):
						yield useForNone
					else:
						yield value

			def getColumn(self, column, *, returnContents = True, trailingNone = False, useForNone = ""):
				"""Returns a list of all openpyxl cell objects within the column
				Any blank cell has a value of None. All columns are the length of the longest column.
				To get just the desired column, make sure that trailingNone is True.

				column (int)       - Which column to retrieve (The first column is '1')
				justLength (bool)  - If True: Returns the number of occupied cells in the column
				contents (bool)    - If True: Returns the cell contents instead of the cell objects
				trailingNone(bool) - If False: Any blank cells will be removed from the end of the list
				useForNone (bool)  - If True: Any internal blank cell will be returned as "" instead of None
									 Note: Only applies if 'contents' is True

				Example Input: getColumn(2)
				Example Input: getColumn("B")
				Example Input: getColumn(2, True)
				Example Input: getColumn(2, trailingNone = False)
				Example Input: getColumn(2, useForNone = None)
				"""

				if (trailingNone):
					_column = self.thing[self.convertColumn(column)]
				else:
					_column = self._filter_trailingNone(self.thing[self.convertColumn(column)])
				
				if (returnContents):
					return tuple(self._yieldContents(_column, useForNone = useForNone))
				return _column

			def getRow(self, row = None, *, returnContents = True, trailingNone = False, useForNone = ""):
				"""Returns a list of all openpyxl cell objects within the row

				row (int)     - Which row to retrieve (The first row is '1')
				justLength (int) - If True: Returns the number of occupied cells in the row
				contents (bool)   - If True: Returns the cell contents instead of the cell objects
				trailingNone(bool)  - If True: Any blank cells will be removed from the end of the list
				noNone (bool)      - If True: Any internal blank cell will be returned as "" instead of None
									 Note: Only applies if 'contents' is True

				Example Input: getRow()
				Example Input: getRow(2)
				Example Input: getRow(2, True)
				Example Input: getRow(2, trailingNone = False)
				Example Input: getRow(2, noNone = False)
				"""

				if (trailingNone):
					_row = self.thing[row]
				else:
					_row = self._filter_trailingNone(self.thing[row])
				
				if (returnContents):
					return tuple(self._yieldContents(_row, useForNone = useForNone))
				return _row

			def getCellWidth(self, row, column):
				"""Returns the [width, height] of a cell.

				row (int)     - The index of the row
				column (int)  - The index of the column. Can be a char

				Example Input: getCellWidth(1, 2)
				"""

				#Retrieve contents
				contents = self.getCellValue(row, column)
				if (contents is not None):
					return len(f"{contents}")

			def setColumnWidth(self, column, newWidth = None):
				"""Changes the width of a column.

				column (int) - The index of the column. Can be a char
				newWidth (int) - The new width of the column. If None: auto adjust the width to the largest value in the column

				Example Input: setColumnWidth(3, 16)
				"""
				def yieldWidths():
					for i, row in enumerate(self.thing.iter_rows(), start = 1):
						width = self.getCellWidth(i, column)
						if (width is not None):
							yield width

				if (newWidth is None):
					#Find the longest cell in the column
					possibleWidths = tuple(yieldWidths())
					if (possibleWidths):
						newWidth = max(possibleWidths)
					else:
						newWidth = -1 #Compensate for blank columns

				#Apply the new width
				newWidth += 2
				self.thing.column_dimensions[openpyxl.utils.get_column_letter(column)].width = newWidth

			def setColumnColor(self, column, color = "CCCCCC"):
				"""Changes the color of a column.

				column (int) - The index of the column. Can be a char
				color (str)    - The new color in hex format

				Example Input: setColumnColor("A")
				"""

				fillObject = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")

				for i, row in enumerate(self.thing.iter_rows(), start = 1):
					cell = self.getCell(row = i , column = column)
					cell.fill = fillObject

if (__name__ == "__main__"):
	excel = Excel()

	with excel.new(label = "test") as myBook:
		with myBook.add(label = "Lorem") as mySheet:
			mySheet.setCell(1, 1, "Ipsum")

		with myBook.add(label = "Dolor") as mySheet:
			mySheet.setCell(1, 2, "Sit")

		myBook.save()
